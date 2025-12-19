import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from ttkthemes import ThemedTk
import threading
import queue
import locale
import os
import roman
import html
import webbrowser
import configparser
import sys
import subprocess

# --- Check and Install webdriver-manager ---
try:
    # Attempt to import to see if it exists
    from webdriver_manager.chrome import ChromeDriverManager
    _WEBDRIVER_MANAGER_INSTALLED = True
except ImportError:
    _WEBDRIVER_MANAGER_INSTALLED = False

def install_webdriver_manager():
    """Attempts to install webdriver-manager using pip."""
    print("Biblioteca 'webdriver-manager' não encontrada. Tentando instalar...")
    try:
        command = [sys.executable, "-m", "pip", "install", "webdriver-manager"]
        result = subprocess.run(command, check=True, capture_output=True, text=True, encoding='utf-8', errors='ignore')
        print("webdriver-manager instalado com sucesso via pip.")
        print(result.stdout)
        from webdriver_manager.chrome import ChromeDriverManager
        return True
    except subprocess.CalledProcessError as e:
        error_message = (
            f"ERRO CRÍTICO: Falha ao instalar 'webdriver-manager' automaticamente.\n"
            f"Verifique sua conexão com a internet e permissões.\n"
            f"Comando executado: {' '.join(e.cmd)}\n"
            f"Código de Saída: {e.returncode}\n"
            f"Output (stderr):\n{e.stderr}\n"
            f"Output (stdout):\n{e.stdout}\n\n"
            f"Por favor, tente instalar manualmente executando no seu terminal:\n"
            f"   {sys.executable} -m pip install webdriver-manager\n\n"
            f"O programa será encerrado."
        )
        print(error_message)
        messagebox.showerror("Erro de Dependência Crítica", error_message)
        return False
    except ImportError:
        error_message = "ERRO CRÍTICO: webdriver-manager foi aparentemente instalado, mas não pôde ser importado. Tente reiniciar o programa ou reinstalar manualmente."
        print(error_message)
        messagebox.showerror("Erro de Dependência Crítica", error_message)
        return False
    except Exception as e:
        error_message = f"ERRO CRÍTICO: Ocorreu um erro inesperado durante a instalação de 'webdriver-manager':\n{e}\n\nO programa será encerrado."
        print(error_message)
        messagebox.showerror("Erro de Dependência Crítica", error_message)
        return False

# --- Proceed only if installation is confirmed ---
if not _WEBDRIVER_MANAGER_INSTALLED:
    if not install_webdriver_manager():
        sys.exit(1)
    try:
        from webdriver_manager.chrome import ChromeDriverManager
        _WEBDRIVER_MANAGER_INSTALLED = True
    except ImportError:
        messagebox.showerror("Erro de Importação", "Falha ao importar webdriver-manager mesmo após a instalação.")
        sys.exit(1)

# Import other necessary modules
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException
import time
import pandas as pd
import fitz
import re
import openpyxl

# --- Constants from legacy script (or slightly adapted) ---
MONTHS = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
MONTH_NUMBERS_MAP = {
    'Jan': '01', 'Fev': '02', 'Mar': '03', 'Abr': '04', 'Mai': '05', 'Jun': '06',
    'Jul': '07', 'Ago': '08', 'Set': '09', 'Out': '10', 'Nov': '11', 'Dez': '12'
}
MONTH_NAMES_MAP_REVERSE = {v: k for k, v in MONTH_NUMBERS_MAP.items()} # For pivot table headers

VENCIMENTO_CODES = ["1101"]
REFERENCIA_REGEX = r'Referência:\s*(\d{4})'
CARGO_REGEX = r'Cargo:\s*([A-Z])\w*(?:\s*-\s*|\s*)([IVX]+)'
TOTAL_PROVENTOS_TEXT = "TOTAL PROVENTOS"
EMPTY_VALUE_MSG = "Valor vazio encontrado para {}. Pulando para o próximo mês."
TOTAL_PROVENTOS_NOT_FOUND_MSG = "TOTAL PROVENTOS não encontrado na página {}"
ERROR_OPENING_EXCEL_FILE_MSG = "Erro ao abrir arquivo Excel para o ano {}: {}"
MISSING_YEARS = {
    1998: 1997, 1999: 1997, 2004: 2003,
    2007: 2006, 2008: 2006, 2011: 2010
}
SELENIUM_TIMEOUT = 15
ORGÃO_RHNET = "309"

CONFIG_FILE = 'config.ini'

def get_config_path():
    """Reads the Excel file path from config.ini, creating a default if it doesn't exist."""
    config = configparser.ConfigParser()
    
    if not os.path.exists(CONFIG_FILE):
        # Create a default config.ini if one does not exist
        print(f"Arquivo de configuração '{CONFIG_FILE}' não encontrado. Criando um modelo...")
        config['Paths'] = {
            '# Instruções': 'Por favor, insira o caminho completo para o arquivo Excel de Vencimentos do Magistério abaixo.',
            'excel_file_path': 'C:/Caminho/Para/VENCIMENTOS MAGISTÉRIO_1993-2014.xlsx'
        }
        with open(CONFIG_FILE, 'w', encoding='utf-8') as configfile:
            config.write(configfile)
        messagebox.showerror(
            "Arquivo de Configuração Criado",
            f"O arquivo de configuração '{CONFIG_FILE}' foi criado no mesmo diretório do programa.\n\n"
            f"Por favor, edite este arquivo para apontar para o local correto do seu arquivo Excel e reinicie o programa."
        )
        return None

    try:
        config.read(CONFIG_FILE, encoding='utf-8')
        path = config.get('Paths', 'excel_file_path')
        return path
    except (configparser.NoSectionError, configparser.NoOptionError) as e:
        messagebox.showerror(
            "Erro no Arquivo de Configuração",
            f"O arquivo '{CONFIG_FILE}' está incompleto ou corrompido.\n"
            f"Ele deve conter uma seção [Paths] com a chave 'excel_file_path'.\n\n"
            f"Erro: {e}\n\n"
            f"Você pode deletar o arquivo '{CONFIG_FILE}' para que um novo modelo seja criado na próxima execução."
        )
        return None
    except Exception as e:
        messagebox.showerror("Erro Inesperado", f"Não foi possível ler o arquivo de configuração: {e}")
        return None

EXCEL_FILE_PATH = get_config_path()

# --- Main Application Class ---

class CalculadoraCHApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Calculadora CH")
        # self.root.geometry("650x550") # Optional: set initial size

        # Center the window
        self.center_window(700, 550)

        try:
            locale.setlocale(locale.LC_NUMERIC, 'pt_BR.UTF-8')
            locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
        except locale.Error as e:
            try:
                locale.setlocale(locale.LC_NUMERIC, 'Portuguese_Brazil.1252')
                locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')
            except locale.Error as e2:
                 locale.setlocale(locale.LC_ALL, '')
                 self.log_message("WARNING", f"Falha ao definir localidade pt_BR: {e} / {e2}. Usando padrão do sistema.")

        # Variables to store user input
        self.login_var = tk.StringVar()
        self.password_var = tk.StringVar()
        self.cpf_var = tk.StringVar()
        self.pdf_path_var = tk.StringVar(value="Nenhum arquivo selecionado")
        self.show_password_var = tk.BooleanVar(value=False)

        # Queue for communication between worker thread and GUI
        self.log_queue = queue.Queue()
        self.result_queue = queue.Queue()

        # Flag to signal cancellation to the worker thread
        self.cancel_requested = threading.Event()

        # Setup GUI elements
        self.create_widgets()

        # Start polling the log queue
        self.root.after(100, self.process_log_queue)

    def center_window(self, width=650, height=550):
        """Centers the Tkinter window on the screen."""
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width/2) - (width/2)
        y = (screen_height/2) - (height/2)
        self.root.geometry('%dx%d+%d+%d' % (width, height, x, y))


    def create_widgets(self):
        """Creates and lays out the GUI widgets."""
        main_frame = ttk.Frame(self.root, padding="10 10 10 10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)

        # --- Credentials Frame ---
        credentials_frame = ttk.LabelFrame(main_frame, text=" Credenciais ", padding="10 10 10 10")
        credentials_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)
        credentials_frame.columnconfigure(1, weight=1)

        ttk.Label(credentials_frame, text="Login RHNet:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.login_entry = ttk.Entry(credentials_frame, textvariable=self.login_var, width=40)
        self.login_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        ttk.Label(credentials_frame, text="Senha RHNet:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.password_entry = ttk.Entry(credentials_frame, textvariable=self.password_var, show="*", width=40)
        self.password_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        self.show_password_check = ttk.Checkbutton(credentials_frame, text="Mostrar Senha",
                                                   variable=self.show_password_var, command=self.toggle_password_visibility)
        self.show_password_check.grid(row=1, column=2, sticky=tk.W, padx=5, pady=5)

        ttk.Label(credentials_frame, text="CPF do Servidor:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.cpf_entry = ttk.Entry(credentials_frame, textvariable=self.cpf_var, width=40)
        self.cpf_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        # --- PDF File Frame ---
        pdf_frame = ttk.LabelFrame(main_frame, text=" Ficha Financeira ", padding="10 10 10 10")
        pdf_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)
        pdf_frame.columnconfigure(1, weight=1)

        self.select_pdf_button = ttk.Button(pdf_frame, text="Selecionar PDF", command=self.select_pdf)
        self.select_pdf_button.grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)

        self.pdf_path_label = ttk.Label(pdf_frame, textvariable=self.pdf_path_var, relief=tk.SUNKEN, anchor=tk.W, width=50)
        self.pdf_path_label.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        # --- Action Buttons Frame ---
        action_frame = ttk.Frame(main_frame, padding="10 0 10 0")
        action_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=10)
        action_frame.columnconfigure(0, weight=1)
        action_frame.columnconfigure(1, weight=1)

        # Style for colored buttons (may depend on theme)
        style = ttk.Style()
        style.configure("Success.TButton", foreground="white", background="green", font=('Helvetica', 10, 'bold'))
        style.map("Success.TButton", background=[('active', 'darkgreen')])
        style.configure("Danger.TButton", foreground="white", background="red", font=('Helvetica', 10, 'bold'))
        style.map("Danger.TButton", background=[('active', 'darkred')])

        self.calculate_button = tk.Button(action_frame, text="CALCULAR",
                                            command=self.start_calculation,
                                            bg="green", fg="white",
                                            font=('Helvetica', 10, 'bold'),
                                            relief=tk.RAISED, borderwidth=2,
                                            activebackground="darkgreen", activeforeground="white")
        self.calculate_button.grid(row=0, column=0, padx=20, pady=5, ipadx=10, ipady=5, sticky=tk.E)

        self.cancel_button = tk.Button(action_frame, text="CANCELAR",
                                        command=self.request_cancel,
                                        bg="red", fg="white",
                                        font=('Helvetica', 10, 'bold'),
                                        relief=tk.RAISED, borderwidth=2,
                                        activebackground="darkred", activeforeground="white")
        self.cancel_button.grid(row=0, column=1, padx=20, pady=5, ipadx=10, ipady=5, sticky=tk.W)
        self.cancel_button.config(state=tk.DISABLED, background='lightgrey', relief=tk.FLAT, disabledforeground='grey40') # Lighter red when disabled

        # --- Log Frame ---
        log_frame = ttk.LabelFrame(main_frame, text=" Log de Eventos ", padding="10 10 10 10")
        log_frame.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)
        main_frame.rowconfigure(3, weight=1)
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        self.log_area = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=10, state=tk.DISABLED)
        self.log_area.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    def toggle_password_visibility(self):
        """Toggles the visibility of the password entry field."""
        if self.show_password_var.get():
            self.password_entry.config(show="")
        else:
            self.password_entry.config(show="*")

    def select_pdf(self):
        """Opens a file dialog to select a PDF file."""
        file_path = filedialog.askopenfilename(
            initialdir="/",
            title="Selecione o arquivo PDF",
            filetypes=(("PDF files", "*.pdf"), ("All files", "*.*"))
        )
        if file_path:
            self.pdf_path_var.set(file_path)
            self.log_message("INFO", f"Ficha Financeira selecionada: {file_path}")
        else:
            self.pdf_path_var.set("Nenhum arquivo selecionado")
            self.log_message("INFO", "Seleção de PDF cancelada.")

    def log_message(self, level, message):
        """Adds a message to the log queue."""
        timestamp = time.strftime("%H:%M:%S")
        log_entry = f"{timestamp} [{level}]: {message}"
        self.log_queue.put(log_entry)

    def process_log_queue(self):
        """Processes messages from the log queue and updates the log area."""
        try:
            while True:
                log_entry = self.log_queue.get_nowait()
                self.log_area.configure(state=tk.NORMAL)
                self.log_area.insert(tk.END, log_entry + "\n")
                self.log_area.configure(state=tk.DISABLED)
                self.log_area.see(tk.END)
                self.log_queue.task_done()
        except queue.Empty:
            self.root.after(100, self.process_log_queue)
        except Exception as e:
             print(f"Erro ao processar fila de log: {e}")
             self.root.after(100, self.process_log_queue)

    def update_gui_state(self, processing=False):
        """Enables/disables widgets based on processing state."""
        gui_state = tk.DISABLED if processing else tk.NORMAL
        cancel_btn_state = tk.NORMAL if processing else tk.DISABLED
        calc_btn_state = tk.DISABLED if processing else tk.NORMAL

        # Standard ttk widgets
        self.login_entry.configure(state=gui_state)
        self.password_entry.configure(state=gui_state)
        self.show_password_check.configure(state=gui_state)
        self.cpf_entry.configure(state=gui_state)
        self.select_pdf_button.configure(state=gui_state)

        # Configure tk.Buttons (CALCULAR)
        self.calculate_button.config(state=calc_btn_state)
        if calc_btn_state == tk.DISABLED:
            self.calculate_button.config(background='lightgrey', relief=tk.FLAT, disabledforeground='grey40')
        else:
            self.calculate_button.config(background='green', relief=tk.RAISED, disabledforeground='lightgrey')

        # Configure tk.Buttons (CANCELAR)
        self.cancel_button.config(state=cancel_btn_state)
        if cancel_btn_state == tk.DISABLED:
            self.cancel_button.config(background='lightgrey', relief=tk.FLAT, disabledforeground='grey40')
        else:
            self.cancel_button.config(background='red', relief=tk.RAISED, disabledforeground='lightgrey')

    def request_cancel(self):
            """Signals the worker thread to cancel the operation."""
            if self.calculate_button['state'] == tk.DISABLED:
                if messagebox.askyesno("Cancelar Operação", "Tem certeza que deseja cancelar o cálculo em andamento?"):
                    self.log_message("WARNING", "Cancelamento solicitado pelo usuário.")
                    self.cancel_requested.set()
                    self.cancel_button.config(state=tk.DISABLED)
            else:
                self.log_message("INFO", "Botão Cancelar clicado, mas nenhum processo ativo.")

    def check_cancel(self):
        """Utility for worker thread to check if cancellation was requested."""
        if self.cancel_requested.is_set():
             self.log_message("INFO", "Processo de cálculo cancelado.")
             return True
        return False

    def start_calculation(self):
        """Validates inputs and starts the calculation in a separate thread."""
        login = self.login_var.get().strip()
        password = self.password_var.get()
        cpf = self.cpf_var.get().strip()
        pdf_path = self.pdf_path_var.get()

        if not all([login, password, cpf]):
            messagebox.showerror("Erro de Entrada", "Login, Senha e CPF são obrigatórios.")
            return
        if not pdf_path or pdf_path == "Nenhum arquivo selecionado":
            messagebox.showerror("Erro de Entrada", "Selecione um arquivo PDF.")
            return
        if not os.path.exists(pdf_path):
             messagebox.showerror("Erro de Arquivo", f"O arquivo PDF selecionado não foi encontrado:\n{pdf_path}")
             return
        if not os.path.exists(EXCEL_FILE_PATH):
             messagebox.showerror("Erro de Arquivo", f"O arquivo Excel de vencimentos não foi encontrado:\n{EXCEL_FILE_PATH}")
             self.log_message("ERROR", f"Arquivo Excel não encontrado em: {EXCEL_FILE_PATH}")
             return

        # Clear previous cancellation request
        self.cancel_requested.clear()

        # Disable GUI elements and enable cancel button
        self.update_gui_state(processing=True)
        self.log_message("INFO", "Iniciando processo de cálculo...")

        # Clear previous results queue
        while not self.result_queue.empty():
            try: self.result_queue.get_nowait()
            except queue.Empty: break

        # Start the worker thread
        self.worker_thread = threading.Thread(
            target=self.run_calculation_thread,
            args=(login, password, cpf, pdf_path),
            daemon=True
        )
        self.worker_thread.start()

        # Start polling the result queue
        self.root.after(200, self.check_calculation_result)

    def check_calculation_result(self):
        """Checks the result queue for completion or errors from the worker thread."""
        try:
            result = self.result_queue.get_nowait()
            reset_gui = True

            if isinstance(result, Exception):
                self.log_message("ERROR", f"Erro durante o cálculo: {result}")
                messagebox.showerror("Erro no Cálculo", f"Ocorreu um erro:\n{result}")
            elif isinstance(result, str) and result == "CANCELLED":
                 self.log_message("INFO", "Processo de cálculo foi cancelado.")
                 messagebox.showwarning("Cancelado", "O processo de cálculo foi cancelado.")
            elif isinstance(result, str) and result == "SUCCESS":
                 self.log_message("INFO", "Processo de cálculo concluído com sucesso.")
                 messagebox.showinfo("Sucesso", "O cálculo foi concluído e a Tabela de CH foi salva.")
            else:
                 self.log_message("WARNING", f"Resultado inesperado do cálculo: {result}")
                 messagebox.showwarning("Atenção", f"Resultado inesperado do processo:\n{result}")

            # Reset GUI state if needed
            if reset_gui:
                 self.update_gui_state(processing=False)

            # Clear cancel flag *after* processing result, ready for next run
            self.cancel_requested.clear()

        except queue.Empty:
            if self.worker_thread.is_alive():
                 self.root.after(200, self.check_calculation_result)
            else:
                 self.log_message("WARNING", "Thread finalizada, mas fila de resultados vazia.")
                 messagebox.showerror("Erro Interno", "A thread de cálculo terminou inesperadamente sem um resultado.")
                 self.update_gui_state(processing=False)
                 self.cancel_requested.clear()

        except Exception as e:
            self.log_message("ERROR", f"Erro ao verificar resultado do cálculo: {e}")
            messagebox.showerror("Erro Interno", f"Erro ao processar resultado:\n{e}")
            self.update_gui_state(processing=False)
            self.cancel_requested.clear()

    # ==================================================================
    # == Core Logic Functions ==
    # ==================================================================

    def run_calculation_thread(self, username, password, cpf, pdf_file_path):
        """The function that runs in the worker thread."""
        driver = None
        operation_status = "UNKNOWN"
        try:
            # --- 1. Parse PDF ---
            if self.check_cancel(): operation_status = "CANCELLED"; return
            self.log_message("INFO", "Analisando PDF...")
            data1 = self.parse_pdf(pdf_file_path)
            if self.cancel_requested.is_set(): operation_status = "CANCELLED"; return
            if data1 is None:
                operation_status = "ERROR"
                return

            # --- 2. Scrape RHNet ---
            if self.check_cancel(): operation_status = "CANCELLED"; return
            self.log_message("INFO", "Acessando RHNet e buscando dados...")
            driver, scraped_data = self.scrape_rhnet(username, password, cpf)
            if self.cancel_requested.is_set(): operation_status = "CANCELLED"; return
            if scraped_data is None:
                driver = None
                operation_status = "ERROR/CANCELLED by scrape_rhnet"
                return

            data2 = scraped_data['data']
            server_info = scraped_data['info']

            # --- 3. Consolidate Data ---
            if self.check_cancel(): operation_status = "CANCELLED"; return
            df1 = pd.DataFrame(data1)
            df2 = pd.DataFrame(data2)
            df = pd.concat([df1, df2]).reset_index(drop=True)

            # Extract months and years from the dates
            try:
                df['Month'] = df['Date'].str[:2].str.strip()
                df['Year'] = df['Date'].str[3:].str.strip()
            except AttributeError as e:
                 self.log_message("ERROR", f"Erro ao extrair Mês/Ano da coluna 'Date'. Verifique os dados. {e}")
                 self.result_queue.put(Exception(f"Formato de data inválido: {e}"))
                 return

            # Get unique years sorted numerically if possible
            try:
                unique_years = sorted(df['Year'].astype(int).unique())
                unique_years_str = [str(y) for y in unique_years]
            except ValueError:
                 unique_years_str = sorted(df['Year'].unique())
                 self.log_message("WARNING", f"Anos não numéricos encontrados: {unique_years_str}. Ordenando como texto.")

            # Use the correct month names for columns from the constant map
            month_columns = list(MONTH_NAMES_MAP_REVERSE.values()) # JAN, FEV, etc.

            # Create an empty pivot table DataFrame
            pivot_table = pd.DataFrame('', index=unique_years_str, columns=month_columns)

            # Fill the pivot table
            for index, row in df.iterrows():
                 year = row['Year']
                 month_num = row['Month']
                 number = row['Number']
                 if year in pivot_table.index:
                     month_name = MONTH_NAMES_MAP_REVERSE.get(month_num)
                     if month_name in pivot_table.columns:
                        current_val = pivot_table.at[year, month_name]
                        if current_val != '' and current_val != str(number):
                             self.log_message("WARNING", f"Múltiplos valores para {month_name}/{year}. Usando último: {number} (anterior: {current_val})")
                        pivot_table.at[year, month_name] = number
                     else:
                         self.log_message("WARNING", f"Número de mês inválido '{month_num}' encontrado para o ano {year}.")
                 else:
                      self.log_message("WARNING", f"Ano '{year}' encontrado nos dados mas não nos anos únicos calculados.")

            if self.check_cancel(): operation_status = "CANCELLED"; return

            # --- 4. Generate HTML ---
            if self.check_cancel(): operation_status = "CANCELLED"; return
            self.log_message("INFO", "Gerando arquivo HTML...")
            self.generate_html(pivot_table, server_info)

        except Exception as e:
            self.log_message("ERROR", f"Erro inesperado na thread de cálculo: {e}")
            self.result_queue.put(e)
            operation_status = "ERROR"
        finally:
            if operation_status == "CANCELLED":
                self.log_message("INFO", "Confirmação de cancelamento na thread.")
                if self.result_queue.empty():
                    self.result_queue.put("CANCELLED")

            # Attempt to close the driver if it exists and wasn't closed already
            if driver:
                try:
                    driver.quit()
                    driver = None
                except Exception as e_quit:
                    self.log_message("WARNING", f"Não foi possível fechar o navegador: {e_quit}")
            else:
                self.log_message("DEBUG", "Nenhuma instância de navegador para fechar.")

    def parse_pdf(self, pdf_file_path):
        """Parses the PDF file to extract ch_number and dates. (Adapted from legacy)"""
        numbers1 = []
        dates1 = []

        try:
            with fitz.open(pdf_file_path) as pdf_document:
                if self.check_cancel(): return None
                self.log_message("INFO", f"PDF contém {pdf_document.page_count} páginas.")
                for page_num in range(pdf_document.page_count):
                    if self.check_cancel(): return None

                    page = pdf_document.load_page(page_num)
                    page_text = page.get_text("text")

                    self.log_message("DEBUG", f"Analisando página {page_num + 1}...")

                    referencia_match = re.search(REFERENCIA_REGEX, page_text)
                    if referencia_match:
                        referencia_year = int(referencia_match.group(1))
                        self.log_message("DEBUG", f"Página {page_num + 1}: Ano de Referência = {referencia_year}")
                        if referencia_year < 1993:
                            self.log_message("INFO", f"Página {page_num + 1}: Ano {referencia_year} < 1993. Pulando página.")
                            continue
                    else:
                        self.log_message("WARNING", f"Referência não encontrada na página {page_num + 1}. Pulando página.")
                        continue

                    cargo_match = re.search(CARGO_REGEX, page_text, re.IGNORECASE)
                    if cargo_match:
                        cargo_text = cargo_match.group(1).upper() + "-" + cargo_match.group(2).upper()
                        self.log_message("DEBUG", f"Página {page_num + 1}: Cargo = {cargo_text}")
                    else:
                        self.log_message("WARNING", f"Cargo não encontrado na página {page_num + 1}. Pulando página.")
                        continue

                    # Find 'VENCIMENTO' amount associated with VENCIMENTO_CODES
                    vencimento_value_str = None
                    vencimento_found = False
                    for vencimento_code in VENCIMENTO_CODES:
                        pattern = re.compile(rf"{vencimento_code}\s+.*?\s+(\d{{1,3}}(?:\.\d{{3}})*(?:,\d{{1,2}}))\b", re.IGNORECASE)
                        match = pattern.search(page_text)
                        if match:
                             vencimento_value_str = match.group(1)
                             vencimento_value_float = locale.atof(vencimento_value_str)
                             self.log_message("DEBUG", f"Página {page_num + 1}: Código {vencimento_code} encontrado. Vencimento Bruto = {vencimento_value_str} ({vencimento_value_float})")
                             vencimento_found = True
                             break

                    if not vencimento_found:
                         self.log_message("DEBUG", f"Página {page_num + 1}: Nenhum código de vencimento {VENCIMENTO_CODES} encontrado com valor numérico. Pulando linha de vencimento.")
                         self.log_message("WARNING", f"Nenhum código de vencimento {VENCIMENTO_CODES} encontrado na página {page_num + 1}. Pulando página.")
                         continue

                    # Find TOTAL PROVENTOS (less critical for CH lookup, but good for context/validation)
                    proventos_match = re.search(rf"{TOTAL_PROVENTOS_TEXT}\s+(\d{{1,3}}(?:\.\d{{3}})*(?:,\d{{1,2}}))\b", page_text, re.IGNORECASE)
                    if proventos_match:
                        total_proventos_str = proventos_match.group(1)
                        self.log_message("DEBUG", f"Página {page_num + 1}: Total Proventos = {total_proventos_str}")
                    else:
                        self.log_message("WARNING", TOTAL_PROVENTOS_NOT_FOUND_MSG.format(page_num + 1))

                    # --- Find corresponding CH in Excel ---
                    excel_year = MISSING_YEARS.get(referencia_year, referencia_year)
                    if excel_year != referencia_year:
                        self.log_message("INFO", f"Dados para {referencia_year} não encontrados no Excel. Usando dados de {excel_year}.")

                    try:
                        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
                        if self.check_cancel(): return None
                        if str(excel_year) not in workbook.sheetnames:
                            self.log_message("ERROR", f"Planilha para o ano {excel_year} não encontrada no arquivo Excel: {EXCEL_FILE_PATH}")
                            continue
                        worksheet = workbook[str(excel_year)]
                        self.log_message("DEBUG", f"Acessando planilha Excel: '{excel_year}'")

                    except FileNotFoundError:
                        self.log_message("ERROR", f"Arquivo Excel não encontrado: {EXCEL_FILE_PATH}")
                        self.result_queue.put(Exception(f"Arquivo Excel não encontrado: {EXCEL_FILE_PATH}"))
                        return None
                    except Exception as e:
                        self.log_message("ERROR", ERROR_OPENING_EXCEL_FILE_MSG.format(excel_year, e))
                        self.result_queue.put(Exception(ERROR_OPENING_EXCEL_FILE_MSG.format(excel_year, e)))
                        return None

                    # Find the relevant month on the PDF page
                    vencimento_associated = []
                    proventos_index = page_text.find(TOTAL_PROVENTOS_TEXT)
                    if proventos_index != -1:      
                        proventos_text_block = page_text[proventos_index + len(TOTAL_PROVENTOS_TEXT):]
                        number_pattern = r'\b\d{1,3}(?:\.\d{3})*(?:,\d{1,2})\b'
                        proventos_numbers_found = re.findall(number_pattern, proventos_text_block)
                        proventos_numbers_on_line = proventos_numbers_found[:12]

                        # Find the VENCIMENTO numbers
                        vencimento_numbers_1101_line = []
                        vencimento_code = VENCIMENTO_CODES[0]
                        vencimento_index = page_text.find(vencimento_code)
                        if vencimento_index != -1:
                            vencimento_text_block = page_text[vencimento_index + len(vencimento_code):]
                            vencimento_numbers_found = re.findall(number_pattern, vencimento_text_block)
                            vencimento_numbers_1101_line = vencimento_numbers_found
                        else:
                            self.log_message("DEBUG", f"Código VENCIMENTO {vencimento_code} não encontrado na página {page_num + 1}.")

                        proventos_numbers_on_line.extend(['0,00'] * (12 - len(proventos_numbers_on_line)))
                        proventos_numbers_on_line = proventos_numbers_on_line[:12]

                        venc_assoc_index = 0
                        for provento_str in proventos_numbers_on_line:
                            try:
                                provento_val = locale.atof(provento_str.strip()) if provento_str else 0.0
                            except ValueError:
                                provento_val = 0.0

                            if provento_val == 0.0:
                                vencimento_associated.append("0")
                            else:
                                if venc_assoc_index < len(vencimento_numbers_1101_line):
                                    venc_num_str = vencimento_numbers_1101_line[venc_assoc_index].strip()
                                    try:
                                        venc_float = locale.atof(venc_num_str)
                                        vencimento_associated.append(f"{venc_float:.2f}")
                                    except ValueError:
                                        self.log_message("WARNING", f"Página {page_num + 1}: Não foi possível converter valor de vencimento '{venc_num_str}' para número.")
                                        vencimento_associated.append("0")
                                    venc_assoc_index += 1
                                else:
                                    self.log_message("WARNING", f"Página {page_num + 1}: Mais valores não-zero em PROVENTOS do que em VENCIMENTO ({vencimento_code}). Faltando dados?")
                                    vencimento_associated.append("0")

                        # Ensure vencimento_associated has 12 entries, padding with "0" if needed
                        while len(vencimento_associated) < 12:
                            vencimento_associated.append("0")
                        vencimento_associated = vencimento_associated[:12]
                        self.log_message("DEBUG", f"Página {page_num + 1}: Vencimento Associado (por mês, baseado em proventos!=0): {vencimento_associated}")

                    else:
                        self.log_message("WARNING", TOTAL_PROVENTOS_NOT_FOUND_MSG.format(page_num + 1))
                        continue
    
                    # --- Loop through months based on legacy logic ---
                    exit_loop_condition_met = False
                    for month_idx, current_month_abbr in enumerate(MONTHS):
                        if self.check_cancel(): return None

                        vencimento_str_for_month = vencimento_associated[month_idx]
                        vencimento_float_for_month = float(vencimento_str_for_month)

                        if vencimento_float_for_month != 0.0:
                            self.log_message("DEBUG", f"Processando Mês: {current_month_abbr} (Venc. Associado: {vencimento_float_for_month})")

                            # Determine the month to use for Excel lookup (Dec for missing years)
                            lookup_month_excel = MONTHS[-1] if referencia_year in MISSING_YEARS else current_month_abbr
                            self.log_message("DEBUG", f"Mês para busca no Excel: {lookup_month_excel} (Cargo: {cargo_text}, Vencimento Base: {vencimento_float_for_month})")

                            # Find the CH Number in the Excel sheet for this specific month's value
                            ch_number = self.find_ch_in_excel(worksheet, lookup_month_excel, cargo_text, vencimento_float_for_month)

                            if ch_number is not None:
                                self.log_message("INFO", f"Página {page_num + 1} ({current_month_abbr}/{referencia_year}): CH encontrado = {ch_number}")
                                numbers1.append(str(ch_number))
                                month_number = MONTH_NUMBERS_MAP[current_month_abbr]
                                dates1.append(f"{month_number}/{referencia_year}")

                                # Check for the specific exit condition from legacy code
                                if referencia_year == 2014 and current_month_abbr == 'Mar':
                                    self.log_message("INFO", "Condição de parada (Mar/2014) atingida na análise do PDF.")
                                    exit_loop_condition_met = True
                                    break
                            else:
                                pass
                        else:
                            pass

                    if exit_loop_condition_met:
                        break

            self.log_message("INFO", "Análise do PDF concluída.")
            return {'Number': numbers1, 'Date': dates1}

        except fitz.fitz.FileNotFoundError:
            self.log_message("ERROR", f"Arquivo PDF não encontrado: {pdf_file_path}")
            self.result_queue.put(Exception(f"Arquivo PDF não encontrado: {pdf_file_path}"))
            return None
        except Exception as e:
            self.log_message("ERROR", f"Erro inesperado ao processar PDF: {e}")
            import traceback
            self.log_message("ERROR", traceback.format_exc())
            self.result_queue.put(Exception(f"Erro ao processar PDF: {e}"))
            return None

    def find_ch_in_excel(self, worksheet, lookup_month, target_cargo, target_vencimento):
        """Finds the CH number in the Excel sheet matching month, cargo, and closest value."""
        self.log_message("DEBUG", f"Buscando no Excel: Mês='{lookup_month}', Cargo='{target_cargo}', Vencimento Base={target_vencimento}")
        DIFFERENCE_THRESHOLD = 5.0
        month_col_index = -1
        cargo_col_index = -1
        ch_col_index = 3
        header_row = 1
        cargo_label_row = 2
        for col_idx in range(1, worksheet.max_column + 1):
            month_cell_value = worksheet.cell(row=header_row, column=col_idx).value
            cargo_cell_value = worksheet.cell(row=cargo_label_row, column=col_idx).value

            if isinstance(month_cell_value, str) and month_cell_value.strip().upper() == lookup_month.upper():
                month_col_index = col_idx
                if isinstance(cargo_cell_value, str) and cargo_cell_value.strip().upper() == 'CARGO':
                     cargo_col_index = col_idx
                     break

        if month_col_index == -1 or cargo_col_index == -1:
             month_row_idx = None
             cargo_label_row_idx = None
             target_cargo_row_idx = None

             for row_idx in range(1, worksheet.max_row + 1):
                 cell_b_val = worksheet.cell(row=row_idx, column=2).value
                 if isinstance(cell_b_val, str):
                    if month_row_idx is None and cell_b_val.strip().upper() == lookup_month.upper():
                        month_row_idx = row_idx
                        self.log_message("DEBUG", f"Mês '{lookup_month}' encontrado na linha {row_idx}, Col B.")
                        for cargo_search_row in range(month_row_idx + 1, worksheet.max_row + 1):
                            sub_cell_b_val = worksheet.cell(row=cargo_search_row, column=2).value
                            if isinstance(sub_cell_b_val, str):
                                 if sub_cell_b_val.strip().upper() in [m.upper() for m in MONTHS]:
                                      self.log_message("DEBUG", f"Encontrado outro mês '{sub_cell_b_val}' antes de achar cargo '{target_cargo}'.")
                                      break
                                 if sub_cell_b_val.strip().upper() == target_cargo.upper():
                                      target_cargo_row_idx = cargo_search_row
                                      self.log_message("DEBUG", f"Cargo '{target_cargo}' encontrado na linha {target_cargo_row_idx}, Col B.")
                                      break

                        if target_cargo_row_idx: break

             if not target_cargo_row_idx:
                 self.log_message("WARNING", f"Não foi possível encontrar a linha para Mês='{lookup_month}' e Cargo='{target_cargo}' na Coluna B do Excel.")
                 return None

             value_col_index = 3
             ch_col_index = 3
             vencimento_col_index = 4 # Col D
             vencimento_cell_value = worksheet.cell(row=target_cargo_row_idx, column=vencimento_col_index).value
             closest_row_idx = None
             closest_distance = float('inf')
             ch_number_for_closest = None
             found_valid_value = False

             if target_cargo_row_idx:
                 rows_to_check = [target_cargo_row_idx]
                 rows_to_check.extend([target_cargo_row_idx + 1, target_cargo_row_idx + 2])

                 for row_idx in rows_to_check:
                      if self.check_cancel(): return None
                      for col_idx in range(vencimento_col_index, worksheet.max_column + 1):
                           if self.check_cancel(): return None
                           cell_value = worksheet.cell(row=row_idx, column=col_idx).value

                           if cell_value is None or cell_value == "-":
                                continue

                           try:
                               if isinstance(cell_value, str):
                                    cell_value_float = locale.atof(cell_value.strip())
                               else:
                                    cell_value_float = float(cell_value)

                               distance = abs(target_vencimento - cell_value_float)

                               if distance < closest_distance:
                                   closest_distance = distance
                                   closest_row_idx = row_idx
                                   ch_number_for_closest = worksheet.cell(row=closest_row_idx, column=ch_col_index).value
                                   found_valid_value = True
                                   if distance < 0.01:
                                       break

                           except (ValueError, TypeError) as e:
                                self.log_message("DEBUG", f"Ignorando valor não numérico '{cell_value}' na célula [{row_idx},{col_idx}]: {e}")
                                continue
                      if closest_distance < 0.01:
                           break
                            
                # --- Store results of the initial search ---
                 initial_ch_number = None
                 initial_distance = closest_distance
                 initial_found = found_valid_value

                 if initial_found:
                     # Validate the initially found CH number
                     if isinstance(ch_number_for_closest, (int, float)):
                         initial_ch_number = ch_number_for_closest
                     elif isinstance(ch_number_for_closest, str) and ch_number_for_closest.strip().isdigit():
                         initial_ch_number = int(ch_number_for_closest)
                     else:
                         self.log_message("WARNING", f"Valor inicial encontrado para CH (Cargo: {target_cargo}) não é numérico: '{ch_number_for_closest}'. Tratando como não encontrado.")
                         initial_found = False

                 # --- Check if re-search is needed ---
                 if initial_found and initial_distance > DIFFERENCE_THRESHOLD:
                     self.log_message("WARNING", f"Diferença inicial ({initial_distance:.2f}) para Cargo '{target_cargo}' excede o limite ({DIFFERENCE_THRESHOLD}). Verificando cargo anterior.")
 
                     # Determine previous cargo
                     previous_cargo = None
                     try:
                         parts = target_cargo.split('-')
                         if len(parts) == 2 and parts[0].upper() == 'P':
                             roman_part = parts[1].upper()
                             if roman_part != 'I':
                                 current_level = roman.fromRoman(roman_part)
                                 if current_level > 1:
                                     previous_level = current_level - 1
                                     previous_cargo = f"P-{roman.toRoman(previous_level)}"
                                     self.log_message("INFO", f"Cargo anterior determinado: {previous_cargo}")
                     except Exception as e:
                         self.log_message("WARNING", f"Não foi possível determinar cargo anterior para '{target_cargo}': {e}")
 
                     if previous_cargo:
                         # --- Perform the second search using previous_cargo ---
                         self.log_message("INFO", f"Realizando nova busca no Excel para Cargo '{previous_cargo}'...")
                         prev_target_cargo_row_idx = None
                         for row_idx in range(1, worksheet.max_row + 1):
                             cell_b_val = worksheet.cell(row=row_idx, column=2).value # Col B = 2
                             if isinstance(cell_b_val, str):
                                 if row_idx > month_row_idx:
                                     if cell_b_val.strip().upper() in [m.upper() for m in MONTHS]:
                                         break
                                     if cell_b_val.strip().upper() == previous_cargo.upper():
                                         prev_target_cargo_row_idx = row_idx
                                         self.log_message("DEBUG", f"Cargo anterior '{previous_cargo}' encontrado na linha {prev_target_cargo_row_idx}, Col B.")
                                         break 
 
                         if prev_target_cargo_row_idx:
                             prev_closest_distance = float('inf')
                             prev_ch_number_for_closest = None
                             prev_found_valid_value = False
                             prev_closest_row_idx = None
                             prev_rows_to_check = [prev_target_cargo_row_idx]
                             prev_rows_to_check.extend([prev_target_cargo_row_idx + 1, prev_target_cargo_row_idx + 2])
 
                             for row_idx in prev_rows_to_check:
                                 if self.check_cancel(): return None
                                 for col_idx in range(vencimento_col_index, worksheet.max_column + 1):
                                     if self.check_cancel(): return None
                                     cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                                     if cell_value is None or cell_value == "-": continue
                                     try:
                                         if isinstance(cell_value, str): cell_value_float = locale.atof(cell_value.strip())
                                         else: cell_value_float = float(cell_value)
 
                                         distance = abs(target_vencimento - cell_value_float)
                                         # self.log_message("DEBUG", f"[Re-Search] Comparando Vencimento Base {target_vencimento} com Excel[{row_idx},{col_idx}] = {cell_value_float} (Dist: {distance})")
 
                                         if distance < prev_closest_distance:
                                             prev_closest_distance = distance
                                             prev_closest_row_idx = row_idx
                                             prev_ch_number_for_closest = worksheet.cell(row=prev_closest_row_idx, column=ch_col_index).value
                                             prev_found_valid_value = True
                                             # self.log_message("DEBUG", f"[Re-Search] Novo valor mais próximo: {cell_value_float} na linha {prev_closest_row_idx}. CH={prev_ch_number_for_closest}")
                                             if distance < 0.01: break
 
                                     except (ValueError, TypeError): continue
                                 if prev_closest_distance < 0.01: break
 
                             # --- Compare results of initial and second search ---
                             if prev_found_valid_value:
                                 prev_ch_number_validated = None
                                 if isinstance(prev_ch_number_for_closest, (int, float)):
                                     prev_ch_number_validated = prev_ch_number_for_closest
                                 elif isinstance(prev_ch_number_for_closest, str) and prev_ch_number_for_closest.strip().isdigit():
                                     prev_ch_number_validated = int(prev_ch_number_for_closest)
 
                                 if prev_ch_number_validated is not None and prev_closest_distance < initial_distance:
                                     self.log_message("INFO", f"Utilizando resultado da re-busca com Cargo '{previous_cargo}'. Distância: {prev_closest_distance:.2f} (CH: {prev_ch_number_validated}).")
                                     return prev_ch_number_validated
                                 else:
                                     self.log_message("INFO", f"Re-busca com Cargo '{previous_cargo}' não produziu resultado melhor (Dist: {prev_closest_distance:.2f}). Mantendo resultado inicial (Dist: {initial_distance:.2f}).")
                                     return initial_ch_number
                             else:
                                 self.log_message("INFO", f"Nenhum valor válido encontrado na re-busca com Cargo '{previous_cargo}'. Mantendo resultado inicial.")
                                 return initial_ch_number
                         else:
                             self.log_message("INFO", f"Linha para Cargo anterior '{previous_cargo}' não encontrada. Mantendo resultado inicial.")
                             return initial_ch_number
 
                     else:
                         self.log_message("WARNING", f"Mantendo resultado inicial para Cargo '{target_cargo}' apesar da alta diferença ({initial_distance:.2f}). Não foi possível/necessário re-buscar cargo anterior.")
                         return initial_ch_number
                 elif initial_found:
                     self.log_message("DEBUG", f"CH {initial_ch_number} encontrado para Cargo '{target_cargo}' com distância aceitável ({initial_distance:.2f}).")
                     return initial_ch_number
                 else:
                     self.log_message("WARNING", f"Nenhum valor de vencimento correspondente encontrado no Excel para Mês='{lookup_month}', Cargo='{target_cargo}', Vencimento Base={target_vencimento}.")
                     return None
 
             else:
                 self.log_message("WARNING", "Lógica falhou em encontrar target_cargo_row_idx, apesar de month_row_idx ter sido encontrado.")
                 return None
 
        self.log_message("WARNING", "Não foi possível determinar o número CH no Excel com a lógica atual (falha inicial na busca de mês/cargo).")
        return None


    def scrape_rhnet(self, username, password, cpf):
        """Logs into RHNet, navigates, and scrapes financial data. (Adapted from legacy)"""
        driver = None
        scraped_data = {'Number': [], 'Date': []}
        server_info = {'nome': 'N/A', 'cargo': 'N/A', 'referencia': 'N/A'}

        try:
            options = webdriver.ChromeOptions()
            options.add_argument("--headless")
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--window-size=1920,1080")
            options.add_argument('log-level=3')
            options.add_experimental_option('excludeSwitches', ['enable-logging'])

            # Use ChromeDriverManager to automatically handle chromedriver
            self.log_message("DEBUG", "Verificando/Instalando chromedriver compatível...")
            try:
                suggested_path = ChromeDriverManager().install()

                driver_path = suggested_path
                expected_exe_name = "chromedriver.exe"
                if not suggested_path.lower().endswith(expected_exe_name.lower()):
                    self.log_message("WARNING", f"Path do webdriver-manager ('{os.path.basename(suggested_path)}') não parece ser o executável ('{expected_exe_name}'). Tentando corrigir...")
                    driver_dir = os.path.dirname(suggested_path)
                    corrected_path = os.path.join(driver_dir, expected_exe_name)
                    if os.path.exists(corrected_path):
                        self.log_message("INFO", f"Usando path corrigido: {corrected_path}")
                        driver_path = corrected_path
                    else:
                        self.log_message("ERROR", f"Path corrigido '{corrected_path}' não encontrado. Usando path original.")
                        driver_path = suggested_path
                
                self.log_message("DEBUG", f"Tentando configurar Service com executable_path: {driver_path}")
                service = Service(executable_path=driver_path)
                self.log_message("DEBUG", "Service configurado.")

                driver = webdriver.Chrome(service=service, options=options)
                self.log_message("DEBUG", "Instância do WebDriver criada com sucesso.")
                driver.implicitly_wait(5)

            except OSError as e:
                 self.log_message("ERROR", f"Erro de Sistema ao obter/usar chromedriver: {e}")
                 if isinstance(e, FileNotFoundError):
                     self.log_message("ERROR", f"O arquivo chromedriver '{driver_path}' não foi encontrado.")
                 elif isinstance(e, PermissionError):
                      self.log_message("ERROR", f"Sem permissão para executar chromedriver '{driver_path}'.")
                 elif "[WinError 193]" in str(e):
                      self.log_message("ERROR", f"O arquivo '{driver_path}' não é um executável válido (WinError 193). Verifique o cache .wdm ou atualize webdriver-manager.")
                 else:
                      self.log_message("ERROR", f"Erro OS não específico: {e}")

                 self.log_message("ERROR", "Verifique também se o Chrome está instalado e atualizado.")
                 self.result_queue.put(Exception(f"Falha ao iniciar ChromeDriver (OSError): {e}"))
                 return None, None
            except Exception as e_manager:
                 self.log_message("ERROR", f"Erro ao inicializar webdriver-manager ou Service: {e_manager}")
                 import traceback
                 self.log_message("ERROR", traceback.format_exc())
                 self.result_queue.put(Exception(f"Falha ao iniciar ChromeDriver: {e_manager}"))
                 return None, None

            except OSError as e:
                self.log_message("ERROR", f"Erro de Sistema ao obter/usar chromedriver: {e}")
                self.log_message("ERROR", "Verifique se há problemas no cache do webdriver-manager (~/.wdm) ou permissões.")
                self.result_queue.put(Exception(f"Falha ao iniciar ChromeDriver (OSError): {e}"))
                return None, None
            except Exception as e_manager:
                self.log_message("ERROR", f"Erro ao inicializar webdriver-manager ou Service: {e_manager}")
                self.result_queue.put(Exception(f"Falha ao iniciar ChromeDriver: {e_manager}"))
                return None, None

            driver.get("https://aplicacoes.expresso.go.gov.br")

            # --- Login ---
            if self.check_cancel(): return None, None
            WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.presence_of_element_located((By.ID, "usernameUserInput"))).send_keys(username)
            WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.presence_of_element_located((By.ID, "password"))).send_keys(password)
            WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.element_to_be_clickable((By.XPATH, '//button[@type="submit"]'))).click()

            # --- Navigation ---
            if self.check_cancel(): driver.quit(); return None, None
            time.sleep(1)

            # Wait for and click the 'people' icon
            # Updated XPath to find the element by the text "RHNet"
            rhnet_xpath = "//h3[normalize-space()='RHNet']"

            # Wait for it to be clickable
            rhnet_link = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, rhnet_xpath)))
            rhnet_link.click()
            time.sleep(2)

            if self.check_cancel(): driver.quit(); return None, None
            WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "menu")))

            # Hover over and click 'Processamento' (using ActionChains)
            processamento_button = WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[2]/div[3]'))) # Adjust XPath if needed
            actions = ActionChains(driver).move_to_element(processamento_button)
            actions.click().perform()
            time.sleep(1)

            # Switch back to default content, then to 'principal' frame
            driver.switch_to.default_content()
            WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "principal")))

            # Hover over and click 'Consultar Ficha Financeira'
            consultar_ficha_button = WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.visibility_of_element_located((By.XPATH, '//div[contains(text(), "Consultar Ficha Financeira")]')))
            ActionChains(driver).move_to_element(consultar_ficha_button).click().perform()
            time.sleep(1)

            # Hover over and click 'Servidor'
            servidor_button = WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.visibility_of_element_located((By.XPATH, '//div[text()="Servidor"]')))
            ActionChains(driver).move_to_element(servidor_button).click().perform()
            time.sleep(1)

            # --- Fill Search Form ---
            if self.check_cancel(): driver.quit(); return None, None
            orgao_xpath = '/html/body/form/center[1]/table/tbody/tr[1]/td[2]/input[2]'
            orgao_textbox = WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.presence_of_element_located((By.XPATH, orgao_xpath)))
            orgao_textbox.send_keys(ORGÃO_RHNET)
            time.sleep(1)

            # CPF textbox
            cpf_xpath = '/html/body/form/center[1]/table/tbody/tr[2]/td[2]/input'
            cpf_textbox = WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.presence_of_element_located((By.XPATH, cpf_xpath)))
            cpf_textbox.send_keys(cpf)
            time.sleep(1)

            # First Dropdown (Tipo Vínculo) - select by index 1 (second option)
            dropdown1_xpath = '/html/body/form/center[1]/table/tbody/tr[3]/td[2]/select'
            select1 = Select(WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
            select1.select_by_index(1)
            time.sleep(2)

            # Second Dropdown (Matrícula) - select by index 1 (second option)
            dropdown2_xpath = '/html/body/form/center[1]/table/tbody/tr[4]/td[2]/select'
            WebDriverWait(driver, SELENIUM_TIMEOUT).until(lambda d: len(Select(d.find_element(By.XPATH, dropdown2_xpath)).options) > 1)
            select2 = Select(driver.find_element(By.XPATH, dropdown2_xpath))
            select2.select_by_index(1)
            time.sleep(1)

            # --- Click Consultar ---
            if self.check_cancel(): driver.quit(); return None, None
            consultar_btn_xpath = '/html/body/form/center[2]/input[1]'
            WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.element_to_be_clickable((By.XPATH, consultar_btn_xpath))).click()
            time.sleep(2)

            # --- Select Record and Get Details ---
            if self.check_cancel(): driver.quit(); return None, None
            try:
                # Click checkbox (adjust XPath/ID if needed, 'marca_desmarca' from legacy)
                checkbox_id = 'marca_desmarca'
                WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.element_to_be_clickable((By.ID, checkbox_id))).click()
                time.sleep(0.5)

                # Click 'Detalhar' button
                detalhar_btn_xpath = '/html/body/form/center[3]/input[2]'
                WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.element_to_be_clickable((By.XPATH, detalhar_btn_xpath))).click()
                time.sleep(2)

            except (TimeoutException, NoSuchElementException) as e:
                self.log_message("ERROR", f"Não foi possível selecionar ou detalhar o registro do servidor: {e}. Verifique o CPF ou se há registros.")
                driver.quit()
                self.result_queue.put(Exception(f"Registro não encontrado/selecionável para CPF {cpf}."))
                return None, None

            # --- Extract Server Info (Nome, Cargo, Referência) ---
            if self.check_cancel(): driver.quit(); return None, None
            try:
                WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/center[1]/table/tbody/tr[4]/td[2]'))) # Nome element

                nome_element = driver.find_element(By.XPATH, '/html/body/form/center[1]/table/tbody/tr[4]/td[2]')
                cargo_element = driver.find_element(By.XPATH, '/html/body/form/center[1]/table/tbody/tr[5]/td[2]')
                referencia_element = driver.find_element(By.XPATH, '/html/body/form/center[1]/table/tbody/tr[6]/td[2]')

                server_info['nome'] = nome_element.text.strip()
                server_info['cargo'] = cargo_element.text.strip()
                server_info['referencia'] = referencia_element.text.strip()

                self.log_message("INFO", f"Nome: {server_info['nome']}")
                self.log_message("INFO", f"Cargo: {server_info['cargo']}")
                self.log_message("INFO", f"Referência: {server_info['referencia']}")

            except (TimeoutException, NoSuchElementException) as e:
                self.log_message("WARNING", f"Não foi possível extrair informações detalhadas do servidor (Nome/Cargo/Ref): {e}")

            # --- Scrape Historical Data (Iteratively click "Recuar") ---
            page_count = 0
            max_pages = 300

            while page_count < max_pages:
                page_count += 1
                if self.check_cancel(): driver.quit(); return None, None

                # Extract Date
                date_text = "N/A"
                try:
                    date_xpath = '/html/body/form/center[1]/table/tbody/tr[1]/td[4]'
                    date_element = WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.visibility_of_element_located((By.XPATH, date_xpath)))
                    date_text = date_element.text.strip()
                    if not re.match(r"\d{2}/\d{4}", date_text):
                         self.log_message("WARNING", f"Formato de data inesperado na pág {page_count}: '{date_text}'. Tentando continuar.")
                except (TimeoutException, NoSuchElementException):
                    self.log_message("WARNING", f"Não foi possível encontrar a data na página {page_count}.")

                # Extract VENCIMENTO EFETIVO Number
                number_text = ""
                try:
                    # Locate the cell with the text "VENCIMENTO EFETIVO" (case-insensitive search might be safer)
                    venc_label_xpath = '//td[contains(translate(text(), "ABCDEFGHIJKLMNOPQRSTUVWXYZ", "abcdefghijklmnopqrstuvwxyz"), "vencimento efetivo")]'
                    vencimento_efetivo_cell = WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.presence_of_element_located((By.XPATH, venc_label_xpath)))

                    # Get the number located in the cell directly to the right
                    next_cell = vencimento_efetivo_cell.find_element(By.XPATH, './following-sibling::td[1]')
                    number_text_raw = next_cell.text.strip()

                    # Clean the number (remove R$, convert comma decimal to dot)
                    number_text_raw = number_text_raw.replace("R$", "").strip()
                    if number_text_raw:
                        try:
                            number_float = locale.atof(number_text_raw)
                            number_text = f"{number_float:.2f}".replace('.', ',')
                            number_text = f"{number_float:.2f}"
                            self.log_message("DEBUG", f"Pág {page_count} ({date_text}): Vencimento Efetivo = {number_text_raw} -> {number_text}")
                        except ValueError:
                            self.log_message("WARNING", f"Pág {page_count} ({date_text}): Não foi possível converter Vencimento Efetivo '{number_text_raw}' para número.")
                            number_text = ""
                    else:
                        number_text = ""

                except (TimeoutException, NoSuchElementException):
                    self.log_message("DEBUG", f"Pág {page_count} ({date_text}): 'VENCIMENTO EFETIVO' não encontrado ou valor adjacente ausente.")
                    number_text = "" # Append empty string as per legacy logic

                # Store extracted data
                if date_text != "N/A":
                    scraped_data['Number'].append(number_text)
                    scraped_data['Date'].append(date_text)
                else:
                    self.log_message("WARNING", f"Pág {page_count}: Ignorando registro devido à data ausente.")


                # Attempt to click "Recuar"
                try:
                    recuar_xpath = '/html/body/form/center[3]/input[1]'
                    recuar_button = WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.element_to_be_clickable((By.XPATH, recuar_xpath)))
                    if recuar_button.is_enabled():
                         recuar_button.click()
                         try:
                             WebDriverWait(driver, SELENIUM_TIMEOUT).until(EC.staleness_of(recuar_button))
                         except TimeoutException:
                             self.log_message("WARNING", f"Pág {page_count}: Botão 'Recuar' não ficou obsoleto após clique. A página pode não ter atualizado.")
                             time.sleep(1)
                    else:
                         self.log_message("INFO", f"Pág {page_count}: Botão 'Recuar' está desabilitado. Fim do histórico alcançado.")
                         break

                except (TimeoutException, NoSuchElementException):
                    self.log_message("INFO", f"Pág {page_count}: Botão 'Recuar' não encontrado ou clicável. Assumindo fim do histórico.")
                    break 
                except Exception as e:
                     self.log_message("ERROR", f"Erro inesperado ao clicar/esperar 'Recuar' na pág {page_count}: {e}")
                     break


            if page_count >= max_pages:
                 self.log_message("WARNING", f"Atingido limite máximo de páginas ({max_pages}) ao clicar em 'Recuar'.")

            self.log_message("INFO", "Extração do RHNet concluída")

            return driver, {'data': scraped_data, 'info': server_info}

        except WebDriverException as e:
            self.log_message("ERROR", f"Erro de WebDriver: {e}")
            if "net::ERR_CONNECTION_REFUSED" in str(e) or "page crash" in str(e):
                 self.log_message("ERROR", "Verifique se o navegador está instalado/atualizado ou se a página está acessível.")
            self.result_queue.put(Exception(f"Erro de WebDriver: {e}"))
            if driver: driver.quit()
            return None, None
        except TimeoutException as e:
            self.log_message("ERROR", f"Tempo limite excedido esperando por elemento: {e.msg}")
            self.result_queue.put(Exception(f"Tempo limite excedido: {e.msg}"))
            if driver: driver.quit()
            return None, None
        except Exception as e:
            self.log_message("ERROR", f"Erro inesperado durante scraping: {e}")
            import traceback
            self.log_message("ERROR", traceback.format_exc())
            self.result_queue.put(Exception(f"Erro inesperado no scraping: {e}"))
            if driver: driver.quit()
            return None, None

    def generate_html(self, pivot_table, server_info):
        """Generates the HTML output file."""

        html_file_path = None
        try:
            # Schedule the file dialog in the main thread
            path_queue = queue.Queue()
            self.root.after(0, lambda: path_queue.put(filedialog.asksaveasfilename(
                 defaultextension=".html",
                 filetypes=(("HTML files", "*.html"), ("All files", "*.*")),
                 title="Salvar Tabela de Cálculo CH",
                 initialfile=f"Calculo_CH_{server_info.get('nome', 'Servidor').replace(' ','_')}.html"
             )))
            # Wait for the result from the main thread
            html_file_path = path_queue.get(timeout=120)

            if not html_file_path:
                self.log_message("WARNING", "Nenhum local selecionado para salvar o arquivo HTML. Geração cancelada.")
                self.result_queue.put("CANCELLED")
                return

            self.log_message("INFO", f"Salvando tabela de CH em: {html_file_path}")

            title = "CÁLCULO DA MÉDIA DE CARGA HORÁRIA ANUAL"
            nome_safe = html.escape(server_info.get('nome', 'N/A'))
            cargo_safe = html.escape(server_info.get('cargo', 'N/A'))
            ref_safe = html.escape(server_info.get('referencia', 'N/A'))
            title2 = f"NOME: {nome_safe}<br>CARGO: {cargo_safe}<br>REFERENCIA: {ref_safe}"

            with open(html_file_path, 'w', encoding='utf-8') as f:
                f.write('<!DOCTYPE html>\n<html lang="pt-BR">\n<head>\n')
                f.write('<meta charset="UTF-8">\n')
                f.write('<meta name="viewport" content="width=device-width, initial-scale=1.0">\n')
                f.write('<title>Cálculo CH</title>\n')
                f.write('<style>\n')
                f.write('  body { font-family: sans-serif; margin: 20px; }\n')
                f.write('  h1 { text-align: center; color: #333; }\n')
                f.write('  h3 { color: #555; border-bottom: 1px solid #ccc; padding-bottom: 10px; margin-bottom: 20px; }\n')
                f.write('  table { border-collapse: collapse; width: 100%; font-size: 12px; text-align: center; margin-top: 15px; }\n')
                f.write('  th, td { border: 1px solid #ccc; padding: 6px 8px; }\n')
                f.write('  th { background-color: #f2f2f2; font-weight: bold; }\n')
                f.write('  td.year-header { font-weight: bold; background-color: #f8f8f8; text-align: center; }\n')
                f.write('  tr:nth-child(even) { background-color: #fafafa; }\n')
                f.write('</style>\n</head>\n<body>\n')

                f.write(f'<h1>{title}</h1>\n')
                f.write(f'<h3>{title2}</h3>\n')

                f.write('<table>\n')
                f.write('<thead>\n<tr><th>Ano</th>')
                for month in pivot_table.columns:
                    f.write(f'<th>{month}</th>')
                f.write('</tr>\n</thead>\n')

                f.write('<tbody>\n')
                for year in pivot_table.index:
                    f.write(f'<tr><td class="year-header">{year}</td>')
                    for month in pivot_table.columns:
                        value_str = str(pivot_table.loc[year, month]).strip()
                        value_display = ""
                        if value_str:
                            try:
                                value_float = float(value_str)
                                value_display = str(int(value_float))
                            except ValueError:
                                if value_str.isdigit():
                                    value_display = value_str
                                else:
                                    value_display = value_str if value_str == "-" else "" 

                        value_safe = html.escape(value_display)
                        f.write(f'<td>{value_safe}</td>')
                    f.write('</tr>\n')
                f.write('</tbody>\n')

                f.write('</table>\n')
                f.write('</body>\n</html>\n')

            self.log_message("INFO", "Tabela de CH gerada com sucesso.")
            try:
                import pathlib
                file_uri = pathlib.Path(html_file_path).as_uri()
                webbrowser.open(file_uri)
            except Exception as e_open:
                self.log_message("WARNING", f"Não foi possível abrir o arquivo HTML automaticamente: {e_open}")

            self.result_queue.put("SUCCESS")

        except queue.Empty:
             self.log_message("ERROR", "Tempo limite excedido esperando pela seleção do local para salvar o arquivo HTML.")
             self.result_queue.put("CANCELLED")
        except Exception as e:
            self.log_message("ERROR", f"Erro ao gerar ou salvar arquivo HTML: {e}")
            import traceback
            self.log_message("ERROR", traceback.format_exc())
            self.result_queue.put(Exception(f"Erro ao gerar/salvar HTML: {e}"))

# --- Main execution ---
if __name__ == "__main__":
    if EXCEL_FILE_PATH is None:
        sys.exit(1)
    try:
        root = ThemedTk(theme="vista")
    except Exception:
        root = tk.Tk()

    app = CalculadoraCHApp(root)
    root.mainloop()